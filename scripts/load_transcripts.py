#!/usr/bin/env python3
"""Load Bolna call transcripts from Google Sheet CSV export and/or local xlsx files."""
import argparse
import csv
import json
import os
import subprocess
import sys
from pathlib import Path


def fetch_google_sheet(sheet_id: str, sheet_name: str) -> list[dict]:
    url = (
        f"https://docs.google.com/spreadsheets/d/{sheet_id}/gviz/tq?"
        f"tqx=out:csv&sheet={sheet_name}"
    )
    try:
        result = subprocess.run(
            ["curl", "-sL", "--max-time", "30", url],
            capture_output=True,
            text=True,
            check=True,
        )
        text = result.stdout
        if not text or "<HTML" in text[:200].upper():
            return []
        reader = csv.DictReader(text.splitlines())
        records = []
        for row in reader:
            records.append(normalize_record(row))
        return [r for r in records if r.get("transcript")]
    except Exception as e:
        print(f"warn: google sheet fetch failed: {e}", file=sys.stderr)
        return []


def normalize_record(row: dict) -> dict:
    lower = {
        str(k).lower().strip(): str(v).strip() if v is not None else ""
        for k, v in row.items()
        if k is not None
    }
    return {
        "timestamp": lower.get("timestamp")
        or lower.get("date")
        or lower.get("created_at")
        or "",
        "lead_id": lower.get("lead_id") or lower.get("lead id") or lower.get("id") or "",
        "direction": lower.get("direction") or lower.get("call_type") or "",
        "transcript": lower.get("transcript") or lower.get("conversation") or "",
        "summary": lower.get("summary") or lower.get("call_summary") or "",
    }


def load_xlsx_dir(input_dir: str) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("warn: openpyxl not installed; skipping xlsx", file=sys.stderr)
        return []
    records = []
    for xlsx in Path(input_dir).glob("*.xlsx"):
        try:
            wb = load_workbook(xlsx, data_only=True, read_only=True)
            for ws in wb.worksheets:
                rows = list(ws.iter_rows(values_only=True))
                if not rows:
                    continue
                headers = [str(h or "").strip() for h in rows[0]]
                for row in rows[1:]:
                    d = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
                    rec = normalize_record(d)
                    if rec.get("transcript"):
                        records.append(rec)
        except Exception as e:
            print(f"warn: failed to read {xlsx}: {e}", file=sys.stderr)
    return records


def combine(gs_path: str, xlsx_path: str, el_path: str = "") -> dict:
    gs = json.load(open(gs_path)) if gs_path and os.path.exists(gs_path) else []
    xl = json.load(open(xlsx_path)) if xlsx_path and os.path.exists(xlsx_path) else []
    el = json.load(open(el_path)) if el_path and os.path.exists(el_path) else []
    seen = set()
    combined = []
    dupes = 0
    for rec in gs + xl + el:
        key = (rec.get("timestamp", ""), rec.get("transcript", "")[:120])
        if key in seen:
            dupes += 1
            continue
        seen.add(key)
        combined.append(rec)
    return {
        "records": combined,
        "stats": {
            "gs": len(gs),
            "xlsx": len(xl),
            "elevenlabs": len(el),
            "dupes": dupes,
            "total": len(combined),
        },
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--source", required=True, choices=["google_sheet", "xlsx", "combine"])
    p.add_argument("--sheet-id")
    p.add_argument("--sheet-name", default="Sheet1")
    p.add_argument("--input-dir")
    p.add_argument("--gs-json")
    p.add_argument("--xlsx-json")
    p.add_argument("--el-json", help="Optional path to ElevenLabs records JSON")
    p.add_argument("--output", required=True)
    args = p.parse_args()

    if args.source == "google_sheet":
        records = fetch_google_sheet(args.sheet_id, args.sheet_name)
        json.dump(records, open(args.output, "w"), ensure_ascii=False, indent=2)
        print(json.dumps({"count": len(records), "output": args.output}))
    elif args.source == "xlsx":
        records = load_xlsx_dir(args.input_dir)
        json.dump(records, open(args.output, "w"), ensure_ascii=False, indent=2)
        print(json.dumps({"count": len(records), "output": args.output}))
    elif args.source == "combine":
        out = combine(args.gs_json, args.xlsx_json, args.el_json or "")
        json.dump(out["records"], open(args.output, "w"), ensure_ascii=False, indent=2)
        print(json.dumps({"output": args.output, **out["stats"]}))


if __name__ == "__main__":
    main()
